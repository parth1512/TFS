import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 100
ws['A2'] = -50
ws['A3'] = 123456
ws['A4'] = 123.45

wb.save('data.xlsx')
print("data.xlsx created successfully")
